"""
generate_sprint_issues.py
Generates the Integrated Search Issues Excel tracker for a given sprint.
Output: integrated-document-search/Integrated_Search_Issues_Sprint_<SPRINT>.xlsx

Usage: python scripts/generate_sprint_issues.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
SPRINT       = "26.06"
OUTPUT_PATH  = rf"C:\Users\diego.jacome\Repos\qa-documentation\integrated-document-search\Integrated_Search_Issues_Sprint_{SPRINT}.xlsx"

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")   # dark navy
DONE_FILL      = PatternFill("solid", fgColor="E2EFDA")   # soft green
TODO_FILL      = PatternFill("solid", fgColor="FFC000")   # amber
OPEN_FILL      = PatternFill("solid", fgColor="DDEBF7")   # light blue
BLOCKED_FILL   = PatternFill("solid", fgColor="FCE4D6")   # salmon/red
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
NOTE_FILL      = PatternFill("solid", fgColor="FFF2CC")   # light yellow (qa notes)
SUPPL_FILL     = PatternFill("solid", fgColor="F2F2F2")   # light gray (supplemental)

WORKING_FILL   = PatternFill("solid", fgColor="92D050")   # green  (matches original)
NOTWORK_FILL   = PatternFill("solid", fgColor="FFC000")   # amber  (matches original)
PARTIAL_FILL   = PatternFill("solid", fgColor="FFEB9C")   # pale yellow (partial/pending verify)

HIGH_FILL   = PatternFill("solid", fgColor="FFD7D7")
MED_FILL    = PatternFill("solid", fgColor="FFEB9C")
LOW_FILL    = PatternFill("solid", fgColor="D9EAD3")

thin   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(size=11):
    return Font(bold=True, color="FFFFFF", size=size, name="Calibri")

def body_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def wrap(h="left", v="top"):
    return Alignment(wrap_text=True, horizontal=h, vertical=v)

# ── Column definitions ────────────────────────────────────────────────────────
# Col A = empty spacer, data B–R
HEADERS = [
    "Item #", "Jira #", "Category", "Description", "Env",
    "Search Affected", "Supplemental", "Priority", "Status",
    "Workaround / Priority Justification", "Reporter",
    "New Searches OFF", "Synchronization Issue?",
    "Tested QA", "Tested in STG", "Dev Note", "QA Note",
]
# Letter → width
COL_WIDTHS = {
    "A": 5.5,
    "B": 8.0,    # Item #
    "C": 11.0,   # Jira #
    "D": 20.0,   # Category
    "E": 36.0,   # Description
    "F": 12.0,   # Env
    "G": 28.0,   # Search Affected
    "H": 42.0,   # Supplemental
    "I": 10.0,   # Priority
    "J": 14.0,   # Status
    "K": 22.0,   # Workaround
    "L": 12.0,   # Reporter
    "M": 10.0,   # New Searches OFF
    "N": 10.0,   # Sync Issue?
    "O": 14.0,   # Tested QA
    "P": 14.0,   # Tested in STG
    "Q": 40.0,   # Dev Note
    "R": 30.0,   # QA Note
}

# ── Issues data ───────────────────────────────────────────────────────────────
# Columns: item, jira, category, description, env, search_affected, supplemental,
#          priority, status, workaround, reporter, new_searches_off, sync_issue,
#          tested_qa, tested_stg, dev_note, qa_note
#
# Status values  : Done | ToDo | Open | Blocked
# Priority values: High | Medium | Low
# Test values    : Working | Not working | Pending | (None)

issues = [
    # ── Carried from Sprint 26.05 ─────────────────────────────────────────────
    (
        1, None, "Fuzzy Search",
        "Underscore operator is not working",
        "QA1", "Regular Search",
        "When searching by Document Attribute value = __d* it should return any "
        "attribute value where the letter 'd' is in 3rd position with any string after.",
        "Low", "ToDo",
        "No workaround available", "Diego J.",
        None, None,
        "Working", None,
        None, None,
    ),
    (
        2, None, "Regular / Custom Search",
        "No results returned using All as Document Attribute Name and searching by value",
        "Pineapple", "Regular / Custom Search",
        "Using All as Document Attribute Name and searching by value or partial value (%test%)\n"
        "Attribute Name: RegularRegressionSecond\n"
        "All: TestAutomationComplexValue\n"
        "All: %TestAutomation%\n"
        "→ 1 document must be returned",
        "High", "Done",
        None, "Diego F.",
        None, None,
        "Not working", "Not working",
        None,
        "When searching for: All:%TestAutomation%. Two results are returned. "
        "One of them does not exist in the database — ES still shows stale document metadata.",
    ),
    (
        3, None, "Attribute / Custom Search",
        "Selecting only Attribute Value returns everything; selected value not shown in filters panel",
        "Pineapple", "Custom Search",
        'Attribute value (location attribute)\n'
        'Everything returned.\n'
        'Value selected not displayed in filters panel.\n'
        'Assumption: selecting only "Attribute Value" should be treated as if "All" was selected as "Attribute Type".',
        "High", "Done",
        None, "Diego F.",
        None, None,
        "Working", None,
        None, None,
    ),
    (
        4, None, "Projects",
        "Documents within project are not included in generated Project Listing file",
        "Encino", "Project Search",
        "When the Project Listing file is generated it only contains the header row; "
        "the documents are not included.",
        "High", "Done",
        None, "Diego F.",
        None, None,
        "Working", "Listing file not created",
        None, None,
    ),
    (
        5, None, "Search Results",
        "Search results counter is not showing total results count once a filter is applied",
        "QA1", "All",
        None,
        "Low", "Done",
        None, "Diego J.",
        None, None,
        "Working", "Not working",
        None, None,
    ),
    (
        6, None, "Search Results",
        "Filtering in search results is not using sub-string exact match",
        "QA1", "All",
        None,
        "High", "Open",
        None, "Diego F.",
        None, None,
        "Working", "Not working",
        None, None,
    ),
    (
        7, None, "Location Attribute Search",
        "Search by location attributes is not returning results",
        "Staging", "Attribute Search",
        "The following search does not return results:\n"
        "Attribute Type: Agreement #\n"
        "Attribute Value: 104180000",
        "High", "Done",
        None, "Diego J.",
        None, None,
        None, "Working",
        "Issue with Stg and Prod indexes — they were not set up correctly. Indexes are now fixed.",
        None,
    ),
    # ── New issues for Sprint 26.06 — add below this line ────────────────────
    # (8, "DD-XXXX", "Category", "Description", "Env", "Search Affected",
    #  "Supplemental details", "High", "Open", None, "Reporter",
    #  None, None, None, None, None, None),
]

# ── Build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"Sprint {SPRINT}"

# Spacer col A
ws.column_dimensions["A"].width = COL_WIDTHS["A"]

# Header row (row 1, data starts at col B = col index 2)
for offset, header in enumerate(HEADERS):
    col_idx = offset + 2               # B=2, C=3, …, R=18
    col_letter = get_column_letter(col_idx)
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font      = hdr_font()
    cell.fill      = HEADER_FILL
    cell.alignment = wrap("center", "center")
    cell.border    = BORDER
    ws.column_dimensions[col_letter].width = COL_WIDTHS[col_letter]

ws.row_dimensions[1].height = 22

# ── Helper maps ───────────────────────────────────────────────────────────────
PRIO_FILL_MAP = {
    "high":   HIGH_FILL,
    "medium": MED_FILL,
    "low":    LOW_FILL,
}
STATUS_FILL_MAP = {
    "done":    DONE_FILL,
    "todo":    TODO_FILL,
    "open":    OPEN_FILL,
    "blocked": BLOCKED_FILL,
}

def test_fill(val):
    if not val:
        return WHITE_FILL
    v = val.lower()
    if "not working" in v:
        return NOTWORK_FILL
    if "working" in v:
        return WORKING_FILL
    if "pending" in v:
        return PARTIAL_FILL
    return WHITE_FILL

# ── Write data rows ───────────────────────────────────────────────────────────
for row_offset, issue in enumerate(issues):
    row_idx = row_offset + 2

    (item, jira, category, description, env, search_aff, suppl,
     priority, status, workaround, reporter, new_off, sync,
     tested_qa, tested_stg, dev_note, qa_note) = issue

    values = [
        item, jira, category, description, env, search_aff, suppl,
        priority, status, workaround, reporter, new_off, sync,
        tested_qa, tested_stg, dev_note, qa_note,
    ]

    for col_offset, val in enumerate(values):
        col_idx    = col_offset + 2
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border    = BORDER
        cell.alignment = wrap()
        cell.font      = body_font()
        cell.fill      = WHITE_FILL

        # ── Column-specific styling ──────────────────────────────────────────
        if col_letter == "B":   # Item #
            cell.alignment = wrap("center", "center")
            cell.font      = body_font(bold=True)

        elif col_letter == "H":   # Supplemental — subtle gray bg
            cell.fill = SUPPL_FILL
            cell.font = body_font(color="444444", size=9)

        elif col_letter == "I":   # Priority
            cell.fill      = PRIO_FILL_MAP.get(str(priority).lower(), WHITE_FILL) if priority else WHITE_FILL
            cell.font      = body_font(bold=True)
            cell.alignment = wrap("center", "center")

        elif col_letter == "J":   # Status
            cell.fill      = STATUS_FILL_MAP.get(str(status).lower(), WHITE_FILL) if status else WHITE_FILL
            cell.font      = body_font(bold=True)
            cell.alignment = wrap("center", "center")

        elif col_letter in ("O", "P"):   # Tested QA / Tested in STG
            cell.fill      = test_fill(val)
            cell.font      = body_font(bold=(val is not None))
            cell.alignment = wrap("center", "center")

        elif col_letter == "R":   # QA Note
            if val:
                cell.fill = NOTE_FILL
            cell.font = body_font(color="444444", size=9)

        elif col_letter == "Q":   # Dev Note
            cell.font = body_font(color="555555", size=9)

    # Auto row height based on content length
    longest = max((len(str(v)) for v in values if v), default=20)
    ws.row_dimensions[row_idx].height = max(45, min(longest * 0.8, 150))

# ── Freeze panes & auto-filter ────────────────────────────────────────────────
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"B1:R1"

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"\n✓ Saved: {OUTPUT_PATH}")
print(f"  Sprint {SPRINT} | {len(issues)} issues")

# ── Quick status summary ───────────────────────────────────────────────────────
status_counts = {}
open_issues   = []
for iss in issues:
    s = iss[8] or "—"
    status_counts[s] = status_counts.get(s, 0) + 1
    if s.lower() not in ("done",):
        open_issues.append(iss)

print("\n── Status Summary ──────────────────────────────────")
for s, c in sorted(status_counts.items()):
    print(f"  {s:10s}: {c}")
print(f"\n── Open / Pending ({len(open_issues)}) ────────────────────────")
for iss in open_issues:
    print(f"  #{iss[0]:2d}  [{iss[8] or 'No status':10s}]  {iss[3][:60]}")
