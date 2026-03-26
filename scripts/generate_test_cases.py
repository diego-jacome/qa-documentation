import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
AREA_FILLS = {
    "Happy Path":       PatternFill("solid", fgColor="D9EAD3"),
    "Error Handling":   PatternFill("solid", fgColor="F4CCCC"),
    "Re-processing":    PatternFill("solid", fgColor="FCE5CD"),
    "Boundary / Limit": PatternFill("solid", fgColor="FFF2CC"),
    "Regression":       PatternFill("solid", fgColor="D6E4F0"),
}
HIGH_FILL   = PatternFill("solid", fgColor="F4CCCC")
MED_FILL    = PatternFill("solid", fgColor="FCE5CD")
LOW_FILL    = PatternFill("solid", fgColor="D9EAD3")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

# ── Headers ──────────────────────────────────────────────────────────────────
headers = ["#", "Area", "Test Case Title", "Preconditions", "Steps", "Expected Result", "Priority", "Notes"]
col_widths = [5, 24, 42, 38, 58, 45, 10, 40]

ws.append(headers)
for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = h
    cell.font = hdr_font()
    cell.fill = HEADER_FILL
    cell.alignment = wrap_align("center", "center")
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 30

# ── Test Cases Data ───────────────────────────────────────────────────────────
# DD-8958 — Email Ingestion Errors for Large Number of Attachments
# Format: (id, area, title, preconditions, steps, expected_result, priority, notes)

test_cases = [

    # ── Happy Path ──────────────────────────────────────────────────────────
    (1, "Happy Path",
     "Email with exactly 13 attachments is fully imported",
     "Tenant configured with valid email ingestion address.\nEmail contains 13 supported-format files (PDF/DOCX/etc.) within size limit.",
     "1. Send email with 13 attachments to the ingestion address.\n2. Wait for the ingestion job to run.\n3. Navigate to the Inbound Email admin screen.\n4. Locate the email and check its status.\n5. Verify all 13 documents appear in Dynamic Docs.",
     "Email shows Completed status.\nAll 13 documents are imported and visible in Dynamic Docs.\nJob Run Details shows: Emails Processed=1, Emails Imported=1, Emails With Errors=0.",
     "High",
     "Pattern 1 threshold boundary (13 attachments was the minimum failing case in production)."),

    (2, "Happy Path",
     "Email with 19 attachments is fully imported",
     "Tenant configured with valid email ingestion address.\nEmail contains 19 supported-format files within size limit.",
     "1. Send email with 19 attachments to the ingestion address.\n2. Wait for the ingestion job to run.\n3. Navigate to the Inbound Email admin screen.\n4. Check status and document count.",
     "Email shows Completed status.\nAll 19 documents are imported successfully.\nNo errors in Job Run Details.",
     "High",
     "Pattern 1 upper boundary (19 attachments also failed in production)."),

    (3, "Happy Path",
     "Email with exactly 20 attachments is fully imported",
     "Tenant configured with valid email ingestion address.\nEmail contains 20 supported-format files within size limit.",
     "1. Send email with 20 attachments to the ingestion address.\n2. Wait for the ingestion job to run.\n3. Navigate to the Inbound Email admin screen and locate the email.\n4. Verify all 20 documents are imported.",
     "Email appears in the Inbound Email admin screen.\nAll 20 documents are imported successfully.\nNo errors.",
     "High",
     "Pattern 2 threshold boundary (20+ attachments were silently dropped in production)."),

    (4, "Happy Path",
     "Email with 23 attachments is fully imported (replicates SG Interest)",
     "Tenant configured with valid email ingestion address.\nEmail contains 23 supported-format files (replicates SG Interest production case).",
     "1. Send email with 23 attachments to the ingestion address.\n2. Wait for the ingestion job to run.\n3. Navigate to the Inbound Email admin screen.\n4. Verify all 23 documents are imported.",
     "Email appears in the Inbound Email admin screen.\nAll 23 documents are imported successfully.\nNo silent drop.",
     "High",
     "Replicates exact SG Interest production scenario (23 PDFs, never appeared in admin screen)."),

    (5, "Happy Path",
     "Email with 100 attachments is fully imported",
     "Tenant configured with valid email ingestion address.\nEmail contains 100 PDF files (~52 KB each, total ~5.2 MB).",
     "1. Send email with 100 attachments via the DD API endpoint (POST /email) or Outlook.\n2. Wait for job execution.\n3. Verify notification email is received with '100 of 100 files imported'.\n4. Check Job Run Details for 0 errors.",
     "All 100 documents are imported.\nNotification email reads 'Imported 100 of 100 files'.\nJob Run Details: Emails Processed=1, Emails Imported=1, Emails With Errors=0.",
     "High",
     "Already validated in QA via API endpoint on 2026-03-12. Re-validate end-to-end when Staging is available."),

    (6, "Happy Path",
     "Email with mixed file types in high quantity is imported",
     "Tenant configured with valid email ingestion address and import rules.\nEmail contains 15+ files of mixed supported types (PDF, DOCX, XLSX, JPG, TIFF).",
     "1. Compose email with a mix of PDF, DOCX, XLSX, JPG, and TIFF files (15+ total).\n2. Send to ingestion address.\n3. Verify all files are imported and classified per Import Rules.",
     "All files are imported and classified according to Document Type rules.\nNo partial imports or missing file types.",
     "Medium",
     "Validates type handling does not degrade with high attachment count."),

    # ── Error Handling ───────────────────────────────────────────────────────
    (7, "Error Handling",
     "Failed email with 13-19 attachments shows a meaningful error message",
     "A simulated or real failure occurs during processing of an email with 13-19 attachments.\nEmail ingestion service is active.",
     "1. Trigger a processing failure (e.g., unsupported file type, corrupted file, misconfigured rule).\n2. Navigate to Inbound Email admin screen.\n3. Locate the failed email.\n4. Inspect the status and error message field.",
     "Email appears in the Inbound Email admin screen with a failed status.\nA meaningful error message is displayed (not blank).\nError is descriptive enough to identify the failure cause.",
     "High",
     "Pre-fix behavior (Pattern 1): email appeared with failed status but no error message. Verify this is now fixed."),

    (8, "Error Handling",
     "Failed email with 20+ attachments appears in admin screen instead of being silently dropped",
     "A simulated or real failure occurs for an email with 20+ attachments.",
     "1. Send an email with 20+ attachments that is expected to fail (e.g., all unsupported file types).\n2. Wait for job run.\n3. Navigate to Inbound Email admin screen.\n4. Confirm the email record is present.",
     "Email record appears in the Inbound Email admin screen.\nStatus shows failed with a meaningful error message.\nEmail is NOT silently discarded.",
     "High",
     "Pre-fix behavior (Pattern 2): email was never written to the inbound queue. Verify it now appears."),

    # ── Re-processing ─────────────────────────────────────────────────────────
    (9, "Re-processing",
     "Manual re-processing of previously failed email (13-19 attachments) succeeds",
     "A pre-fix failed email with 13-19 attachments exists in the admin screen.",
     "1. Locate a previously failed email (Pattern 1) in the Inbound Email admin screen.\n2. Click the manual re-process action.\n3. Wait for the job to run.\n4. Verify the outcome.",
     "Email is re-processed successfully.\nAll documents from the email are imported into Dynamic Docs.\nStatus updates to Completed with no error.",
     "High",
     "Explicit AC: 'Manual re-processing of previously failed high-attachment emails succeeds after the fix.'"),

    (10, "Re-processing",
     "Manual re-processing of previously failed email (20+ attachments) succeeds",
     "Post-fix: a previously silently-dropped email with 20+ attachments has been added to the queue.",
     "1. Locate a known previously failed/dropped email with 20+ attachments.\n2. Trigger manual re-processing.\n3. Wait for the job run to complete.\n4. Verify all documents are imported.",
     "All documents from the email are imported.\nStatus shows Completed.\nJob Run Details shows 0 errors.",
     "High",
     "Validates that the fix also recovers emails that were silently dropped before the patch."),

    # ── Boundary / Limit ──────────────────────────────────────────────────────
    (11, "Boundary / Limit",
     "Email with 12 attachments processes normally (below bug threshold)",
     "Tenant configured with valid email ingestion address.\nEmail contains 12 supported-format files.",
     "1. Send email with 12 attachments.\n2. Wait for job run.\n3. Verify all 12 documents are imported without errors.",
     "All 12 documents imported successfully.\nNo errors.\nBehavior is unchanged from pre-fix baseline.",
     "Medium",
     "Baseline below the 13-attachment failure threshold. Ensures fix did not alter behavior for smaller emails."),

    (12, "Boundary / Limit",
     "Email exceeding the client size limit (~20 MB) is handled gracefully",
     "Tenant configured with valid email ingestion address.\nEmail total attachment size exceeds 20 MB.",
     "1. Attempt to send an email with total attachment size exceeding 20 MB.\n2. Observe behavior at the email client level and in the admin screen.",
     "Email client rejects the send (enforced by client).\nIf email is received despite size, system shows a meaningful error — it does not crash or drop silently.",
     "Medium",
     "Validates size-limit handling is separate from attachment-count handling."),

    # ── Regression ────────────────────────────────────────────────────────────
    (13, "Regression",
     "Email with 1 attachment continues to process normally",
     "Tenant configured with valid email ingestion address.\nEmail contains 1 supported-format file.",
     "1. Send email with a single attachment.\n2. Wait for job run.\n3. Verify the document is imported.",
     "Document is imported successfully.\nNo regression introduced for standard single-attachment emails.",
     "Medium",
     "Basic regression check to confirm the fix did not break the standard happy path."),

    (14, "Regression",
     "Email with no attachments is handled without error",
     "Tenant configured with valid email ingestion address.\nEmail body has no file attachments.",
     "1. Send an email with no attachments to the ingestion address.\n2. Wait for job run.\n3. Check the Inbound Email admin screen.",
     "Email is processed without crashing the service.\nNo documents are imported.\nStatus is Completed or shows a descriptive informational message.",
     "Low",
     "Ensures the fix for high-attachment emails did not introduce issues for the no-attachment edge case."),

    (15, "Regression",
     "Import Rules are correctly applied when email contains 20+ attachments",
     "Tenant has at least one Import Rule configured (e.g., Subject contains 'Invoice' → Document Type = Invoice).\nEmail with 20+ attachments matching the rule is sent.",
     "1. Send email with subject matching an Import Rule and 20+ attachments.\n2. Wait for job run.\n3. Verify each imported document is classified correctly per the rule.",
     "All documents are imported and classified according to the Import Rule (correct Department, Cabinet, Folder, Document Type).\nRule application is not skipped due to high attachment count.",
     "High",
     "Validates that Import Rules still apply correctly at high attachment counts after the fix."),
]

# ── Write rows ───────────────────────────────────────────────────────────────
for tc in test_cases:
    tc_id, area, title, precond, steps, expected, priority, query_str = tc
    row = [tc_id, area, title, precond, steps, expected, priority, query_str]
    ws.append(row)
    row_idx = ws.max_row

    # Area fill
    area_fill = AREA_FILLS.get(area, WHITE_FILL)

    # Priority fill
    if priority == "High":
        prio_fill = HIGH_FILL
    elif priority == "Medium":
        prio_fill = MED_FILL
    else:
        prio_fill = LOW_FILL

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = wrap_align()
        if col_idx == 7:
            cell.fill = prio_fill
            cell.font = cell_font(bold=True)
        elif col_idx == 8:
            cell.fill = PatternFill("solid", fgColor="F5F5F5")
            cell.font = cell_font(bold=False, color="555555", size=9)
        elif col_idx in (1, 2):
            cell.fill = area_fill
            cell.font = cell_font(bold=(col_idx == 1))
        else:
            cell.fill = area_fill
            cell.font = cell_font(bold=(col_idx == 3))

    ws.row_dimensions[row_idx].height = 120

# ── Freeze panes & filters ───────────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

output_path = r"C:\Users\diego.jacome\Repos\qa-documentation\test-cases\DD-8958\DD-8958_TestCases.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}  ({len(test_cases)} test cases)")
