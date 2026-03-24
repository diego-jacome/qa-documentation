import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
AREA_FILLS = {
    "Sync / Backfill Integrity":          PatternFill("solid", fgColor="D6E4F0"),
    "Results vs. Stored Procedures":      PatternFill("solid", fgColor="D9EAD3"),
    "Multi-tenant Isolation":             PatternFill("solid", fgColor="FCE5CD"),
    "Search Types":                       PatternFill("solid", fgColor="EAD1DC"),
    "Critical Business Filters":          PatternFill("solid", fgColor="FFF2CC"),
    "Nested — Hashtags":                  PatternFill("solid", fgColor="CFE2F3"),
    "Nested — Comments":                  PatternFill("solid", fgColor="D9D2E9"),
    "Nested — Packet / Projects":         PatternFill("solid", fgColor="FCE5CD"),
    "Nested — Attribute Values":          PatternFill("solid", fgColor="D9EAD3"),
    "Nested — linked_to":                 PatternFill("solid", fgColor="FFF2CC"),
    "Nested — Post-Sync / Backfill":      PatternFill("solid", fgColor="F4CCCC"),
    "Pagination & Sorting":               PatternFill("solid", fgColor="D0E0E3"),
    "Edge Cases":                         PatternFill("solid", fgColor="EAD1DC"),
}
HIGH_FILL   = PatternFill("solid", fgColor="F4CCCC")
MED_FILL    = PatternFill("solid", fgColor="FCE5CD")
LOW_FILL    = PatternFill("solid", fgColor="D9EAD3")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def cell_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

# ── Headers ──────────────────────────────────────────────────────────────────
headers = ["#", "Area", "Test Case Title", "Preconditions", "Steps", "Expected Result", "Priority", "Sample ES Query"]
col_widths = [5, 28, 42, 38, 55, 45, 10, 80]

ws.append(headers)
for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = h
    cell.font = hdr_font()
    cell.fill = HEADER_FILL
    cell.alignment = wrap_align("center", "center")
    cell.border = border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 30

# ── Test Cases Data ───────────────────────────────────────────────────────────
# Format: (id, area, title, preconditions, steps, expected_result, priority, sample_query)

SYSTEM_ID = "63"

def q(body: str) -> str:
    return f'GET documents/_search\n{body}'

test_cases = [

    # ── Area 1: Sync / Backfill Integrity ─────────────────────────────────
    (1, "Sync / Backfill Integrity",
     "Pre-backfill documents appear in search",
     "Documents exist in DB created before the backfill date.",
     "1. Pick a known document created before backfill.\n2. Search by doc_name via the Search API.\n3. Verify result appears.",
     "Document is returned in results. Fields match DB values.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"doc_name": "Pre-backfill Doc Name"}}\n      ]\n    }\n  }\n}')),

    (2, "Sync / Backfill Integrity",
     "Post-backfill documents appear in search",
     "Documents exist in DB created after the backfill completed.",
     "1. Pick a known document created after backfill.\n2. Search by content_id via the Search API.\n3. Verify result appears.",
     "Document is returned. Index data matches DB state.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"content_id": 987654}}\n      ]\n    }\n  }\n}')),

    (3, "Sync / Backfill Integrity",
     "Documents modified during backfill reflect final state",
     "A document was modified while backfill was running.",
     "1. Identify a document edited during the backfill window.\n2. Compare ES fields with current DB values.",
     "ES reflects the most recent DB state, no stale data.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"content_id": 111222}}\n      ]\n    }\n  }\n}')),

    (4, "Sync / Backfill Integrity",
     "No duplicate documents in the index",
     "Backfill and sync pipeline have both run.",
     "1. Run aggregation by content_id.\n2. Check if any content_id has more than one doc with is_current=1.",
     "Each content_id has exactly one document with is_current=1.",
     "High",
     q('{\n  "size": 0,\n  "query": {"term": {"system_id": "' + SYSTEM_ID + '"}},\n  "aggs": {\n    "by_content_id": {\n      "terms": {"field": "content_id", "min_doc_count": 2},\n      "aggs": {\n        "current_count": {"filter": {"term": {"is_current": 1}}}\n      }\n    }\n  }\n}')),

    (5, "Sync / Backfill Integrity",
     "Soft-deleted documents do not appear in normal search",
     "At least one document marked as is_soft_deleted=1 exists.",
     "1. Search using softDeletedFilter: ExcludeDeleted.\n2. Verify soft-deleted document is absent from results.",
     "No soft-deleted documents appear in results.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"is_soft_deleted": 0}}\n      ]\n    }\n  }\n}')),

    # ── Area 2: Results vs. Stored Procedures ─────────────────────────────
    (6, "Results vs. Stored Procedures",
     "Search by doc_name matches legacy SP results",
     "Known doc_name exists in DB and was indexed.",
     "1. Run old SP with doc_name filter.\n2. Run ES search with same filter.\n3. Compare result count and document IDs.",
     "Both return the same set of documents.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"doc_name": "Contract 2024"}}\n      ]\n    }\n  }\n}')),

    (7, "Results vs. Stored Procedures",
     "Full-text content search returns relevant results without false positives",
     "Documents with known content text are indexed.",
     "1. Search by a keyword known to exist in content.\n2. Verify all returned docs contain that keyword.\n3. Verify no unrelated docs are returned.",
     "All results contain the keyword. No false positives.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"content": "annual report"}}\n      ]\n    }\n  }\n}')),

    (8, "Results vs. Stored Procedures",
     "Attribute search returns only matching documents (no cross-attribute false positives)",
     "Documents with different attributes exist.",
     "1. Search by attributeId + specific value.\n2. Verify all returned documents have that exact attribute+value combination.",
     "No document is returned where the value belongs to a different attribute.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 5}},\n                {"match": {"attribute_values.value": "Approved"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (9, "Results vs. Stored Procedures",
     "Date range filter returns correct results inclusive of boundaries",
     "Documents with creation_date at range boundaries exist.",
     "1. Search with creationDateFrom = T and creationDateTo = T+n.\n2. Verify documents at exactly T and T+n are included.\n3. Verify documents outside range are excluded.",
     "Boundary dates are included. No out-of-range results.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"range": {"creation_date": {"gte": 1700000000000, "lte": 1710000000000}}}\n      ]\n    }\n  }\n}')),

    (10, "Results vs. Stored Procedures",
     "Location filter returns only documents from that location",
     "Documents with different primary_location_id values exist.",
     "1. Search by primaryLocationId.\n2. Verify all results belong to that location only.",
     "Only documents with the specified primary_location_id are returned.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"primary_location_id": 42}}\n      ]\n    }\n  }\n}')),

    # ── Area 3: Multi-tenant Isolation ─────────────────────────────────────
    (11, "Multi-tenant Isolation",
     "User from system A cannot see documents from system B",
     "Documents exist under two different system_ids.",
     "1. Search with system_id = A.\n2. Confirm no documents with system_id = B appear.",
     "Zero documents from system B in the results.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}}\n      ],\n      "must_not": [\n        {"term": {"system_id": "99"}}\n      ]\n    }\n  }\n}')),

    (12, "Multi-tenant Isolation",
     "Filters applied in one tenant do not affect results from another",
     "Two tenants have documents with the same doc_name.",
     "1. Search by doc_name in tenant A.\n2. Confirm results are scoped to tenant A only.",
     "Results are isolated to the queried system_id.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"doc_name": "Invoice"}}\n      ]\n    }\n  }\n}')),

    # ── Area 4: Search Types ───────────────────────────────────────────────
    (13, "Search Types",
     "Regular search — keyword returns matching documents",
     "Documents with searchable content exist.",
     "1. Use searchType: Regular with a keyword.\n2. Verify results match on doc_name or content.",
     "Results contain the keyword in doc_name or content.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"multi_match": {\n          "query": "invoice",\n          "fields": ["doc_name", "content"]\n        }}\n      ]\n    }\n  }\n}')),

    (14, "Search Types",
     "Attribute/Comment/Hashtag search filters correctly by each independently",
     "Documents with attributes, comments, and hashtags exist.",
     "1. Search by attribute only — verify comment/hashtag fields don't interfere.\n2. Repeat for comment only, then hashtag only.",
     "Each filter type returns only documents matching that specific field.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"match": {"hashtags.name": "legal"}}\n        }}\n      ]\n    }\n  }\n}')),

    (15, "Search Types",
     "Custom search — multiple filters return their intersection",
     "Documents exist that match some but not all filters.",
     "1. Search with attribute + location + date range combined.\n2. Verify only documents matching ALL conditions are returned.",
     "Result is the intersection of all applied filters.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"primary_location_id": 42}},\n        {"range": {"creation_date": {"gte": 1700000000000}}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {"term": {"attribute_values.attribute_id": 5}}\n        }}\n      ]\n    }\n  }\n}')),

    (16, "Search Types",
     "Tree search respects full location hierarchy",
     "Documents exist at different levels of the location hierarchy.",
     "1. Search by tree using a parent location.\n2. Verify documents from child locations are also returned.",
     "Results include documents from the parent and all descendant locations.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"terms": {"primary_location_id": [42, 43, 44]}}\n      ]\n    }\n  }\n}')),

    # ── Area 5: Critical Business Filters ─────────────────────────────────
    (17, "Critical Business Filters",
     "softDeletedFilter: ExcludeDeleted hides all deleted documents",
     "Documents with is_soft_deleted=1 exist.",
     "1. Search with softDeletedFilter: ExcludeDeleted.\n2. Confirm no document with is_soft_deleted=1 appears.",
     "Zero deleted documents in results.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"is_soft_deleted": 0}}\n      ]\n    }\n  }\n}')),

    (18, "Critical Business Filters",
     "softDeletedFilter: IncludeDeleted returns all documents including deleted",
     "Both deleted and non-deleted documents exist.",
     "1. Search with softDeletedFilter: IncludeDeleted.\n2. Confirm deleted documents appear in results.",
     "Results include documents where is_soft_deleted=0 and is_soft_deleted=1.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}}\n      ]\n    }\n  }\n}')),

    (19, "Critical Business Filters",
     "isLocked filter returns only locked / non-locked documents",
     "Both locked and unlocked documents exist.",
     "1. Search with isLocked: true — verify only locked docs returned.\n2. Search with isLocked: false — verify only unlocked docs returned.",
     "Results match the lock state filter exactly.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"is_locked": 1}}\n      ]\n    }\n  }\n}')),

    (20, "Critical Business Filters",
     "isCurrent filter returns only the current version of each document",
     "Multiple versions exist for some documents.",
     "1. Search with isCurrent: true.\n2. For any document with multiple versions, verify only the current one appears.",
     "Only documents with is_current=1 are returned.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"is_current": 1}}\n      ]\n    }\n  }\n}')),

    (21, "Critical Business Filters",
     "processingStatus filter returns documents in the correct state",
     "Documents with various processing statuses exist.",
     "1. Search filtering by processingStatus = 'Processed'.\n2. Verify all returned documents have that status.\n3. Repeat for 'Unprocessed'.",
     "Only documents with the specified status are returned.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"processing_status.keyword": "Processed"}}\n      ]\n    }\n  }\n}')),

    (22, "Critical Business Filters",
     "OCR confidence range filter applies correct boundaries",
     "Documents with varying ocr_confidence values exist.",
     "1. Search with minOCRConfidence=0.7 and maxOCRConfidence=1.0.\n2. Verify all results have ocr_confidence within that range.\n3. Verify documents outside range are excluded.",
     "Results respect min/max boundaries inclusively.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"range": {"ocr_confidence": {"gte": 0.7, "lte": 1.0}}}\n      ]\n    }\n  }\n}')),

    # ── Area 6: Nested — Hashtags ──────────────────────────────────────────
    (23, "Nested — Hashtags",
     "Search by hashtag A returns only documents tagged with hashtag A",
     "Documents with hashtag A and documents with hashtag B exist.",
     "1. Search by hashtag A.\n2. Verify every returned document has hashtag A.\n3. Verify documents with only hashtag B are excluded.",
     "Only documents containing hashtag A are returned.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"term": {"hashtags.name.keyword": "legal"}}\n        }}\n      ]\n    }\n  }\n}')),

    (24, "Nested — Hashtags",
     "Hashtag B text does not match documents that only have a similar hashtag A",
     "Documents with hashtag 'legal' and hashtag 'legal-review' exist.",
     "1. Search by exact hashtag 'legal'.\n2. Confirm 'legal-review' documents are not returned.",
     "No false positives from partial text match.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"term": {"hashtags.name.keyword": "legal"}}\n        }}\n      ]\n    }\n  }\n}')),

    (25, "Nested — Hashtags",
     "Document with multiple hashtags [A, B] is found by searching either independently",
     "A document with hashtags [finance, legal] exists.",
     "1. Search by hashtag 'finance' — verify document appears.\n2. Search by hashtag 'legal' — verify same document appears.",
     "Document is returned in both searches.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"term": {"hashtags.name.keyword": "finance"}}\n        }}\n      ]\n    }\n  }\n}')),

    (26, "Nested — Hashtags",
     "Document without hashtags does not appear in hashtag search",
     "A document with no hashtags exists.",
     "1. Search by any hashtag.\n2. Verify the document with no hashtags is not in results.",
     "Document with empty hashtags array is excluded.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"exists": {"field": "hashtags.name"}}\n        }}\n      ]\n    }\n  }\n}')),

    (27, "Nested — Hashtags",
     "After adding a hashtag, document appears in search for that hashtag",
     "A document without hashtag 'urgent' exists.",
     "1. Add hashtag 'urgent' to the document.\n2. Wait for sync.\n3. Search by hashtag 'urgent'.\n4. Verify document now appears.",
     "Document is returned after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"term": {"hashtags.name.keyword": "urgent"}}\n        }}\n      ]\n    }\n  }\n}')),

    (28, "Nested — Hashtags",
     "After removing a hashtag, document no longer appears in that hashtag search",
     "A document with hashtag 'urgent' exists.",
     "1. Remove hashtag 'urgent' from the document.\n2. Wait for sync.\n3. Search by hashtag 'urgent'.\n4. Verify document is gone.",
     "Document is absent from results after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"term": {"hashtags.name.keyword": "urgent"}}\n        }}\n      ]\n    }\n  }\n}')),

    (29, "Nested — Hashtags",
     "Hashtag search is case-insensitive (or behavior is consistent and documented)",
     "A document with hashtag 'Legal' (capitalized) exists.",
     "1. Search by hashtag 'legal' (lowercase).\n2. Document if result appears or not.\n3. Behavior must be consistent across all hashtags.",
     "Behavior is consistent and matches documented expectation.",
     "Low",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "hashtags",\n          "query": {"match": {"hashtags.name": "legal"}}\n        }}\n      ]\n    }\n  }\n}')),

    # ── Area 7: Nested — Comments ──────────────────────────────────────────
    (30, "Nested — Comments",
     "Search by userId in comments returns only documents where that user commented",
     "Documents with comments from multiple users exist.",
     "1. Search by userId=5.\n2. Verify all returned documents have at least one comment from user 5.",
     "No document appears where user 5 has no comments.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {"term": {"comments.user_id": 5}}\n        }}\n      ]\n    }\n  }\n}')),

    (31, "Nested — Comments",
     "Search by comment text returns only documents containing that text in a comment",
     "Documents with various comment texts exist.",
     "1. Search by comment text 'approved for release'.\n2. Verify every returned document has a comment with that text.",
     "No document is returned where the text appears only in content or other fields.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {"match": {"comments.comment": "approved for release"}}\n        }}\n      ]\n    }\n  }\n}')),

    (32, "Nested — Comments",
     "[CRITICAL] userId + comment text must match within the SAME comment object",
     "Document has: Comment 1 from user 5 with text 'pending'. Comment 2 from user 9 with text 'approved'.",
     "1. Search by userId=5 AND comment='approved'.\n2. Verify this document is NOT returned (false positive scenario).",
     "Document is not returned. Conditions evaluated per comment, not across comments.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"comments.user_id": 5}},\n                {"match": {"comments.comment": "approved"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (33, "Nested — Comments",
     "Document with no comments does not appear in comment search",
     "A document with no comments exists.",
     "1. Search by any userId or comment text.\n2. Verify document with no comments is not returned.",
     "Document with empty comments array is excluded.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {"exists": {"field": "comments.comment"}}\n        }}\n      ]\n    }\n  }\n}')),

    (34, "Nested — Comments",
     "After adding a comment, document appears in search for that comment/user",
     "Document without comments from user 7 exists.",
     "1. Add a comment from user 7.\n2. Wait for sync.\n3. Search by userId=7.\n4. Verify document appears.",
     "Document is returned in results after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {"term": {"comments.user_id": 7}}\n        }}\n      ]\n    }\n  }\n}')),

    (35, "Nested — Comments",
     "After deleting the only comment from a user, document no longer appears for that userId",
     "Document has only one comment from user 7.",
     "1. Delete the comment from user 7.\n2. Wait for sync.\n3. Search by userId=7.\n4. Verify document is absent.",
     "Document is excluded from results after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "comments",\n          "query": {"term": {"comments.user_id": 7}}\n        }}\n      ]\n    }\n  }\n}')),

    # ── Area 8: Nested — Packet / Projects ────────────────────────────────
    (36, "Nested — Packet / Projects",
     "Search by packetId returns only documents in that packet",
     "Documents assigned to different packets exist.",
     "1. Search by packetId=100.\n2. Verify all results belong to packet 100.\n3. Verify documents in packet 200 do not appear.",
     "Only documents in the specified packet are returned.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.packet_id": 100}}\n      ]\n    }\n  }\n}')),

    (37, "Nested — Packet / Projects",
     "Document in multiple packets appears in search for each packet",
     "A document assigned to packets [P1, P2] exists.",
     "1. Search by packetId=P1 — verify document appears.\n2. Search by packetId=P2 — verify document appears.\n3. Search by packetId=P99 — verify document does not appear.",
     "Document is returned for P1 and P2 only.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.packet_id": 200}}\n      ]\n    }\n  }\n}')),

    (38, "Nested — Packet / Projects",
     "[CRITICAL] Packet object flattening — packetId + order must not produce false positives",
     "Document in packet P1 with order=1 and packet P2 with order=2.",
     "1. Search for packetId=P1 AND order=2.\n2. Verify document is NOT returned (object flattening false positive scenario).",
     "Document is not returned. If it is returned, confirms the object vs nested bug.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.packet_id": 100}},\n        {"term": {"packet.order": 2}}\n      ]\n    }\n  }\n}')),

    (39, "Nested — Packet / Projects",
     "is_read_only filter on packet returns only read-only assigned documents",
     "Documents with is_read_only=true and false packets exist.",
     "1. Search for documents in a packet where is_read_only=true.\n2. Verify all results are in read-only packets.",
     "Only documents in read-only packets are returned.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.is_read_only": true}}\n      ]\n    }\n  }\n}')),

    (40, "Nested — Packet / Projects",
     "After adding a document to a project, it appears in packet search",
     "Document not in packet P1 exists.",
     "1. Add document to packet P1.\n2. Wait for sync.\n3. Search by packetId=P1.\n4. Verify document appears.",
     "Document is returned in packet P1 search after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.packet_id": 150}}\n      ]\n    }\n  }\n}')),

    (41, "Nested — Packet / Projects",
     "After removing a document from a project, it no longer appears in that packet search",
     "Document in packet P1 exists.",
     "1. Remove document from packet P1.\n2. Wait for sync.\n3. Search by packetId=P1.\n4. Verify document is absent.",
     "Document is not returned for packet P1 after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"packet.packet_id": 150}}\n      ]\n    }\n  }\n}')),

    # ── Area 9: Nested — Attribute Values ─────────────────────────────────
    (42, "Nested — Attribute Values",
     "Attribute search returns only documents where the exact attributeId has the exact value",
     "Documents with different attributes and values exist.",
     "1. Search by attributeId=5 with value='Approved'.\n2. Verify all results have attribute 5 = 'Approved'.\n3. Verify documents where attribute 5 has another value are excluded.",
     "Only exact attributeId + value matches are returned.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 5}},\n                {"term": {"attribute_values.value.keyword": "Approved"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (43, "Nested — Attribute Values",
     "[CRITICAL] Cross-attribute false positive — attributeId A with value Y must not match",
     "Document has: Attribute 5 with value 'Pending'. Attribute 8 with value 'Approved'.",
     "1. Search by attributeId=5 AND value='Approved'.\n2. Verify this document is NOT returned.",
     "Document is not returned. No cross-attribute match.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 5}},\n                {"term": {"attribute_values.value.keyword": "Approved"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (44, "Nested — Attribute Values",
     "Numeric range filter on attribute evaluates conditions within the same attribute",
     "Documents with numeric attributes exist. Document has attr 3=10, attr 4=50.",
     "1. Search by attributeId=3 with numberFrom=40, numberTo=60.\n2. Verify document is NOT returned (10 is out of range for attr 3).",
     "No false positive from range condition mixing across attributes.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 3}},\n                {"range": {"attribute_values.date_value": {"gte": 40, "lte": 60}}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (45, "Nested — Attribute Values",
     "Date range filter on attribute evaluates conditions within the same attribute",
     "Document has attr 6 with date T1, attr 7 with date T2.",
     "1. Search by attributeId=6 with dateFrom=T2.\n2. Verify document is NOT returned (attr 6 has T1, not within range).",
     "No false positive from mixing date ranges across attributes.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 6}},\n                {"range": {"attribute_values.date_value": {"gte": 1710000000000}}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (46, "Nested — Attribute Values",
     "Document with multiple attributes of same type — each evaluated independently",
     "Document has two AlphaNumeric attributes: 5='Active', 6='Inactive'.",
     "1. Search by attributeId=5 value='Inactive'.\n2. Verify document is NOT returned.",
     "Attribute conditions are evaluated per attribute object, not merged.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 5}},\n                {"term": {"attribute_values.value.keyword": "Inactive"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (47, "Nested — Attribute Values",
     "After modifying an attribute value, search returns the updated value",
     "Document has attributeId=5 with value='Draft'.",
     "1. Update attribute 5 value to 'Published'.\n2. Wait for sync.\n3. Search by attributeId=5 value='Published'.\n4. Verify document appears.\n5. Search by value='Draft' — verify document no longer appears.",
     "Document reflects the updated attribute value after sync.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"attribute_values.attribute_id": 5}},\n                {"term": {"attribute_values.value.keyword": "Published"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    # ── Area 10: Nested — linked_to ────────────────────────────────────────
    (48, "Nested — linked_to",
     "Search by tertiaryId returns only documents linked to that entity",
     "Documents linked to different entities exist.",
     "1. Search by tertiaryId=300.\n2. Verify all results are linked to entity 300.\n3. Verify documents linked only to entity 400 are excluded.",
     "Only documents linked to tertiaryId=300 are returned.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "linked_to",\n          "query": {"term": {"linked_to.tertiary_id": 300}}\n        }}\n      ]\n    }\n  }\n}')),

    (49, "Nested — linked_to",
     "linked_to tertiaryId + name must match within the same link object",
     "Document linked to entity A (id=300, name='Alpha') and entity B (id=400, name='Beta').",
     "1. Search by tertiaryId=300 AND name='Beta'.\n2. Verify document is NOT returned.",
     "No cross-object match between different linked_to entries.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "linked_to",\n          "query": {\n            "bool": {\n              "must": [\n                {"term": {"linked_to.tertiary_id": 300}},\n                {"term": {"linked_to.name.keyword": "Beta"}}\n              ]\n            }\n          }\n        }}\n      ]\n    }\n  }\n}')),

    (50, "Nested — linked_to",
     "After removing a link, document no longer appears for that entity search",
     "Document linked to entity 300.",
     "1. Remove link to entity 300.\n2. Wait for sync.\n3. Search by tertiaryId=300.\n4. Verify document is absent.",
     "Document is excluded after sync.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "linked_to",\n          "query": {"term": {"linked_to.tertiary_id": 300}}\n        }}\n      ]\n    }\n  }\n}')),

    # ── Area 11: Nested — Post-Sync / Backfill ─────────────────────────────
    (51, "Nested — Post-Sync / Backfill",
     "Backfilled documents have all nested arrays correctly populated",
     "Backfill has completed.",
     "1. Pick 5+ documents that existed before backfill and have known attributes, hashtags, and comments in DB.\n2. Query ES and compare nested arrays against DB records.",
     "All nested fields match DB. No empty arrays where data should exist.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"content_id": 655789}}\n      ]\n    }\n  }\n}')),

    (52, "Nested — Post-Sync / Backfill",
     "Documents modified post-backfill reflect updated nested objects with no stale data",
     "A document was modified after backfill (new hashtag, updated attribute).",
     "1. Compare ES nested fields against DB for modified document.\n2. Verify old values are gone and new values are present.",
     "Nested objects reflect current DB state, no stale entries.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"content_id": 112233}}\n      ]\n    }\n  }\n}')),

    (53, "Nested — Post-Sync / Backfill",
     "Document with large nested arrays (10+ attributes, 5+ comments) — all elements searchable",
     "A document with many attributes and comments exists.",
     "1. Search by each attribute individually.\n2. Search by each comment userId.\n3. Verify all nested elements are reachable via search.",
     "Every nested element is individually searchable with no truncation.",
     "High",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"nested": {\n          "path": "attribute_values",\n          "query": {"term": {"attribute_values.attribute_id": 12}}\n        }}\n      ]\n    }\n  }\n}')),

    # ── Area 12: Pagination & Sorting ──────────────────────────────────────
    (54, "Pagination & Sorting",
     "Pagination returns consistent results without duplicates or skips between pages",
     "More than pageSize documents exist for a given query.",
     "1. Request page 1 with size=10.\n2. Request page 2 with size=10.\n3. Verify no document appears on both pages.\n4. Verify no document is skipped.",
     "All pages together cover the complete result set without overlap or gaps.",
     "High",
     q('{\n  "from": 0,\n  "size": 10,\n  "query": {\n    "bool": {\n      "must": [{"term": {"system_id": "' + SYSTEM_ID + '"}}]\n    }\n  }\n}')),

    (55, "Pagination & Sorting",
     "Sort by doc_name uses .keyword subfield for correct lexicographic order",
     "Multiple documents with different doc_name values exist.",
     "1. Search with sort by doc_name ascending.\n2. Verify results are in alphabetical order.",
     "Results are sorted correctly. No case-sensitive ordering issues.",
     "Medium",
     q('{\n  "query": {"term": {"system_id": "' + SYSTEM_ID + '"}},\n  "sort": [{"doc_name.keyword": {"order": "asc"}}]\n}')),

    (56, "Pagination & Sorting",
     "Sort by dynamic attribute (AlphaNumeric type) works correctly",
     "Documents with an AlphaNumeric attribute exist.",
     "1. Search sorted by a specific attributeId of type AlphaNumeric.\n2. Verify results are in the expected order.",
     "Results are sorted by the attribute value in the correct order.",
     "Medium",
     q('{\n  "query": {"term": {"system_id": "' + SYSTEM_ID + '"}},\n  "sort": [\n    {"attribute_values.value.keyword": {\n      "order": "asc",\n      "nested": {\n        "path": "attribute_values",\n        "filter": {"term": {"attribute_values.attribute_id": 5}}\n      }\n    }}\n  ]\n}')),

    # ── Area 13: Edge Cases ────────────────────────────────────────────────
    (57, "Edge Cases",
     "Search with all filters empty returns results without error",
     "Documents exist in the index.",
     "1. Send search request with all filter fields empty/null.\n2. Verify response is 200 with results.",
     "API returns paginated results, no 400 or 500 error.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [{"term": {"system_id": "' + SYSTEM_ID + '"}}]\n    }\n  }\n}')),

    (58, "Edge Cases",
     "Search with a non-existent content_id returns empty results, not an error",
     "The queried content_id does not exist in the index.",
     "1. Search by content_id that does not exist.\n2. Verify response is 200 with 0 results.",
     "Response is 200 with empty hits array.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"term": {"content_id": 9999999999}}\n      ]\n    }\n  }\n}')),

    (59, "Edge Cases",
     "Search with special characters in keyword does not cause query errors",
     "System is running normally.",
     "1. Search with keyword containing: %, *, \", /, <, >.\n2. Verify each returns 200 (with or without results).",
     "No 400 or 500 error. System handles special characters gracefully.",
     "Medium",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"doc_name": "test/document%2024"}}\n      ]\n    }\n  }\n}')),

    (60, "Edge Cases",
     "Documents with empty content (no OCR text) are searchable by other fields",
     "A document with content='' or null exists.",
     "1. Search by doc_name for a document known to have empty content.\n2. Verify the document appears.",
     "Document is returned. Empty content does not prevent searchability.",
     "Low",
     q('{\n  "query": {\n    "bool": {\n      "must": [\n        {"term": {"system_id": "' + SYSTEM_ID + '"}},\n        {"match": {"doc_name": "Scanned Empty Doc"}}\n      ],\n      "must_not": [{"exists": {"field": "content"}}]\n    }\n  }\n}')),
]

# ── Write rows ───────────────────────────────────────────────────────────────
for tc in test_cases:
    tc_id, area, title, precond, steps, expected, priority, query_str = tc
    row = [tc_id, area, title, precond, steps, expected, priority, query_str]
    ws.append(row)
    row_idx = ws.max_row

    # Area fill
    area_fill = AREA_FILLS.get(area, WHITE_FILL)

    # Priority fill
    if priority == "High":
        prio_fill = HIGH_FILL
    elif priority == "Medium":
        prio_fill = MED_FILL
    else:
        prio_fill = LOW_FILL

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = border
        cell.alignment = wrap_align()
        if col_idx == 7:
            cell.fill = prio_fill
            cell.font = cell_font(bold=True)
        elif col_idx == 8:
            cell.fill = PatternFill("solid", fgColor="F5F5F5")
            cell.font = Font(name="Courier New", size=8, color="1F4E79")
        elif col_idx in (1, 2):
            cell.fill = area_fill
            cell.font = cell_font(bold=(col_idx == 1))
        else:
            cell.fill = area_fill
            cell.font = cell_font(bold=(col_idx == 3))

    ws.row_dimensions[row_idx].height = 120

# ── Freeze panes & filters ───────────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

output_path = r"c:\Users\diego.jacome\Repos\qa-documentation\integrated-document-search\ES_Search_TestCases.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}  ({len(test_cases)} test cases)")
